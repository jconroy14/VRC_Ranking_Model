import requests
from openpyxl import Workbook, load_workbook
from rankingModel import getAccuracy

filename = 'rankingModelData.xlsx'
wb = load_workbook(filename)
ws = wb['Data']

# Get events
url = "https://api.vexdb.io/v1/get_events"
params = {'program':'VRC','season':'In The Zone','status':'past','limit_number':100,'limit_start':100}
events = requests.get(url,params).json()

# S-NE Championship sku: RE-VRC-17-4151
#NZ Nationals: RE-VRC-17-4422
# North Texas Regionals: RE-VRC-17-3012
# Kansas:RE-VRC-17-4564
# Ohio: RE-VRC-17-3589
# North NY: RE-VRC-17-2871
# DelMarVa: RE-VRC-18-4291
# Pennsulvania: RE-VRC-17-4533
eventsSkipped = 0
startIndex = 70
for eventIndex in xrange(events['size']):
    sku = events['result'][eventIndex]['sku']

    # define matches parameter
    url = "https://api.vexdb.io/v1/get_matches"
    params = {'sku':sku,'scored':1}
    call = requests.get(url,params).json()
    qualificationMatches = []
    eliminationMatches = []
    for match in call['result']:
        if match['round'] == 2:
            qualificationMatches.append(match)
        elif match['round'] in [3,4,5]:
            eliminationMatches.append(match)
    numMatches = len(eliminationMatches) + len(qualificationMatches)

    if(numMatches>0): #only continue if there is data for the event
        # define ranking parameter
        url = "https://api.vexdb.io/v1/get_rankings"
        params = {'sku':sku}
        rankings = requests.get(url,params).json()

        # define model - combined_rank | ts_ccwm | ccwm | ts | rank | ap | red
        numMatch_model = [0,0,0,1,0,0,0]
        combined_rank_model = [1,0,1,1,1,1,0]
        ts_ccwm_model = [0,1,0,0,0,0,0]
        ccwm_model = [0,0,1,0,0,0,0]
        ts_model = [0,0,0,1,0,0,0]
        rank_model = [0,0,0,0,1,0,0]
        ap_model = [0,0,0,0,0,1,0]
        red_model = [0,0,0,0,0,0,1]

        # eliminate ties
        tieMargin = 0

        # call to get accuracy
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=1).value = events['result'][eventIndex]['name']
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=2).value = getAccuracy(numMatch_model,qualificationMatches,eliminationMatches,rankings,tieMargin)[0]
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=3).value = getAccuracy(combined_rank_model,qualificationMatches,eliminationMatches,rankings,tieMargin)[1]
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=4).value = getAccuracy(ts_ccwm_model,qualificationMatches,eliminationMatches,rankings,tieMargin)[1]
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=5).value = getAccuracy(ccwm_model,qualificationMatches,eliminationMatches,rankings,tieMargin)[1]
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=6).value = getAccuracy(ts_model,qualificationMatches,eliminationMatches,rankings,tieMargin)[1]
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=7).value = getAccuracy(rank_model,qualificationMatches,eliminationMatches,rankings,tieMargin)[1]
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=8).value = getAccuracy(ap_model,qualificationMatches,eliminationMatches,rankings,tieMargin)[1]
        ws.cell(row=eventIndex+startIndex-eventsSkipped,column=9).value = getAccuracy(red_model,qualificationMatches,eliminationMatches,rankings,tieMargin)[1]

    else:
        eventsSkipped +=1 #if event is skipped, mark it so that there are no blank rows in the spreadsheet

wb.save(filename)
